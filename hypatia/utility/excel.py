from hypatia.backend.ModelSettings import ModelSettings
from hypatia.backend.ModelData import ModelData
from hypatia.utility.constants import ModelMode
from openpyxl import load_workbook
import pandas as pd
import os
import shutil
from hypatia.error_log.Exceptions import (
    WrongInputMode,
    DataNotImported,
    ResultOverWrite,
    SolverNotFound,
)

def read_settings(path: str, mode: ModelMode) -> ModelSettings:
    wb_glob = load_workbook(r"{}/global.xlsx".format(path))
    sets_glob = wb_glob["Sets"]

    global_settings = {}
    for entry, data_boundary in sets_glob.tables.items():
        data_glob = sets_glob[data_boundary]
        content = [[cell.value for cell in ent] for ent in data_glob]
        header = content[0]
        rest = content[1:]
        df = pd.DataFrame(rest, columns=header)
        global_settings[entry] = df

    regional_settings = {}
    for reg in list(global_settings["Regions"]["Region"]):
        wb = load_workbook(r"{}/{}.xlsx".format(path, reg))
        sets = wb["Sets"]
        settings = {}
        for entry, data_boundary in sets.tables.items():
            data = sets[data_boundary]
            content = [[cell.value for cell in ent] for ent in data]
            header = content[0]
            rest = content[1:]
            df = pd.DataFrame(rest, columns=header)
            settings[entry] = df

        regional_settings[reg] = settings

    return ModelSettings(mode, global_settings, regional_settings)


def write_parameters_files(settings: ModelSettings, path: str, force_rewrite: bool = False):
    if os.path.exists(path):
        if not force_rewrite:
            raise ResultOverWrite(
                f"Folder {path} already exists. To over write"
                f" the parameter files, use force_rewrite=True."
            )
        else:
            shutil.rmtree(path)
    os.mkdir(path)
    if settings.multi_node:
        with pd.ExcelWriter(
            r"{}/parameters_connections.xlsx".format(path)
        ) as writer:
            for key, value in settings.trade_parameters_template.items():
                connection_data = pd.DataFrame(
                    value["value"], index=value["index"], columns=value["columns"]
                )
                connection_data.to_excel(writer, sheet_name=value["sheet_name"])

        with pd.ExcelWriter(r"{}/parameters_global.xlsx".format(path)) as writer:
            for key, value in settings.global_parameters_template.items():
                global_data = pd.DataFrame(
                    value["value"], index=value["index"], columns=value["columns"]
                )
                global_data.to_excel(writer, sheet_name=value["sheet_name"])

    for reg in settings.regions:
        with pd.ExcelWriter(r"{}/parameters_{}.xlsx".format(path, reg)) as writer:
            for key, value in settings.regional_parameters_template[reg].items():
                regional_data = pd.DataFrame(
                    value["value"], index=value["index"], columns=value["columns"]
                )
                regional_data.to_excel(writer, sheet_name=value["sheet_name"])

def read_parameters(settings: ModelSettings, path: str) -> ModelData:
    global_parameters = None
    if settings.global_parameters_template != None:
        global_parameters = {}
        for key, value in settings.global_parameters_template.items():
            global_parameters[key] = pd.read_excel(
                r"{}/parameters_global.xlsx".format(path),
                sheet_name=value["sheet_name"],
                index_col=list(range(0, __get_number_of_levels(value["index"]))),
                header=list(range(0, __get_number_of_levels(value["columns"]))),
            )

    trade_parameters = None
    if settings.trade_parameters_template != None:
        trade_parameters = {}
        for key, value in settings.trade_parameters_template.items():
            trade_parameters[key] = pd.read_excel(
                r"{}/parameters_connections.xlsx".format(path),
                sheet_name=value["sheet_name"],
                index_col=list(range(0, __get_number_of_levels(value["index"]))),
                header=list(range(0, __get_number_of_levels(value["columns"]))),
            )

    regional_parameters = {}
    for region in settings.regions:
        parameters = {}
        for key, value in settings.regional_parameters_template[region].items():
            parameters[key] = pd.read_excel(
                r"{}/parameters_{}.xlsx".format(path, region),
                sheet_name=value["sheet_name"],
                index_col=list(range(0, __get_number_of_levels(value["index"]))),
                header=list(range(0, __get_number_of_levels(value["columns"]))),
            )
        regional_parameters[region] = parameters

    return ModelData(
        settings,
        global_parameters,
        trade_parameters,
        regional_parameters,
    )

def __get_number_of_levels(index):
    if isinstance(index, pd.MultiIndex):
        return len(index.names)
    return 1
