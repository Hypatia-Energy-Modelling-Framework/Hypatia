# -*- coding: utf-8 -*-
"""
This module contains the ReadSets class that is in charge
of reading the sets files, reshaping them to be used in
the build class, creating and reading the parameter files and
checking the errors in the definition of the sets and parameters
"""

import itertools as it
from openpyxl import load_workbook
import pandas as pd
from hypatia.error_log.Checks import (
    check_nan,
    check_index,
    check_index_data,
    check_table_name,
    check_mapping_values,
    check_mapping_ctgry,
    check_sheet_name,
    check_tech_category,
    check_carrier_type,
    check_years_mode_consistency,
)

from hypatia.error_log.Exceptions import WrongInputMode
import numpy as np
from hypatia.utility.constants import (
    global_set_ids,
    regional_set_ids,
    technology_categories,
    carrier_types,
)
from hypatia.utility.constants import (list_connection_operation, list_connection_planning,
 take_regional_sheets, take_trade_ids, take_ids, take_global_ids)

MODES = ["Planning", "Operation"]


class ReadSets:

    """ Class that reads the sets of the model, creates the parameter files with
    default values and reads the filled parameter files
    
    Attributes
    ------------
    mode: 
        The mode of optimization including the operation and planning mode
        
    path:
        The path of the set files given by the user
        
    glob_mapping : dict
        A dictionary of the global set tables given by the user in the global.xlsx file
    
    mapping : dict
        A dictionary of the regional set tables given by the user in the regional
        set files
    
    connection_sheet_ids:  dict
        A nested dictionary that defines the sheet names of the parameter file of
        the inter-regional links with their default values, indices and columns

    global_sheet_ids : dict
        A nested dictionary that defines the sheet names of the global parameter file
        with their default values, indices and columns
        
    regional_sheets_ids : dict
        A nested dictionary that defines the sheet names of the regional parameter files
        with their default values, indices and columns
        
    trade_data : dict
        A nested dictionary for storing the inter-regional link data
        
    global_data : dict
        A nested dictionary for storing the global  data
        
    data : dict
        A nested dictionary for storing the regional data
    """

    def __init__(self, path, mode="Planning"):

        self.mode = mode
        self.path = path

        self._init_by_xlsx()

    def _init_by_xlsx(self,):

        """
        Reads and organizes the global and regional sets
        """
        glob_mapping = {}
        wb_glob = load_workbook(r"{}/global.xlsx".format(self.path))
        sets_glob = wb_glob["Sets"]
        set_glob_category = {key: value for key, value in sets_glob.tables.items()}

        for entry, data_boundary in sets_glob.tables.items():

            data_glob = sets_glob[data_boundary]
            content = [[cell.value for cell in ent] for ent in data_glob]
            header = content[0]
            rest = content[1:]
            df = pd.DataFrame(rest, columns=header)
            glob_mapping[entry] = df

        self.glob_mapping = glob_mapping

        check_years_mode_consistency(
            mode=self.mode, main_years=list(self.glob_mapping["Years"]["Year"])
        )

        for key, value in self.glob_mapping.items():
            check_table_name(
                file_name="global",
                allowed_names=list(global_set_ids.keys()),
                table_name=key,
            )

            check_index(value.columns, key, "global", pd.Index(global_set_ids[key]))

            check_nan(key, value, "global")

            if key == "Technologies":

                check_tech_category(value, technology_categories, "global")

            if key == "Carriers":

                check_carrier_type(value, carrier_types, "global")

        self.regions = list(self.glob_mapping["Regions"]["Region"])
        self.main_years = list(self.glob_mapping["Years"]["Year"])

        if "Timesteps" in self.glob_mapping.keys():

            self.time_steps = list(self.glob_mapping["Timesteps"]["Timeslice"])
            self.timeslice_fraction = self.glob_mapping["Timesteps"][
                "Timeslice_fraction"
            ].values

        else:

            self.time_steps = ["Annual"]
            self.timeslice_fraction = np.ones((1, 1))

        # possible connections among the regions

        if len(self.regions) > 1:
            lines_obj = it.permutations(self.regions, r=2)

            self.lines_list = []
            for item in lines_obj:

                if item[0] < item[1]:

                    self.lines_list.append("{}-{}".format(item[0], item[1]))

        mapping = {}

        for reg in self.regions:

            wb = load_workbook(r"{}/{}.xlsx".format(self.path, reg))

            sets = wb["Sets"]

            self._setbase_reg = [
                "Technologies",
                "Carriers",
                "Carrier_input",
                "Carrier_output",
            ]

            set_category = {key: value for key, value in sets.tables.items()}

            reg_mapping = {}

            for entry, data_boundary in sets.tables.items():

                data = sets[data_boundary]
                content = [[cell.value for cell in ent] for ent in data]

                header = content[0]

                rest = content[1:]

                df = pd.DataFrame(rest, columns=header)

                reg_mapping[entry] = df

            mapping[reg] = reg_mapping

            for key, value in mapping[reg].items():

                check_table_name(
                    file_name=reg,
                    allowed_names=list(regional_set_ids.keys()),
                    table_name=key,
                )

                check_index(value.columns, key, reg, pd.Index(regional_set_ids[key]))

                check_nan(key, value, reg)

                if key == "Technologies":

                    check_tech_category(value, technology_categories, reg)

                if key == "Carriers":

                    check_carrier_type(value, carrier_types, reg)

                if key == "Carrier_input" or key == "Carrier_output":

                    check_mapping_values(
                        value,
                        key,
                        mapping[reg]["Technologies"],
                        "Technologies",
                        "Technology",
                        "Technology",
                        reg,
                    )

            check_mapping_values(
                mapping[reg]["Carrier_input"],
                "Carrier_input",
                mapping[reg]["Carriers"],
                "Carriers",
                "Carrier_in",
                "Carrier",
                reg,
            )

            check_mapping_values(
                mapping[reg]["Carrier_output"],
                "Carrier_output",
                mapping[reg]["Carriers"],
                "Carriers",
                "Carrier_out",
                "Carrier",
                reg,
            )

            check_mapping_ctgry(
                mapping[reg]["Carrier_input"],
                "Carrier_input",
                mapping[reg]["Technologies"],
                "Supply",
                reg,
            )

            check_mapping_ctgry(
                mapping[reg]["Carrier_output"],
                "Carrier_output",
                mapping[reg]["Technologies"],
                "Demand",
                reg,
            )

        self.mapping = mapping

        Technologies = {}

        for reg in self.regions:

            regional_tech = {}

            for key in list(self.mapping[reg]["Technologies"]["Tech_category"]):

                regional_tech[key] = list(
                    self.mapping[reg]["Technologies"].loc[
                        self.mapping[reg]["Technologies"]["Tech_category"] == key
                    ]["Technology"]
                )

            Technologies[reg] = regional_tech

        self.Technologies = Technologies

        self._create_input_data()

    def _create_input_data(self):
        """
        Defines the sheets, indices and columns of the parameter files
        """

        if len(self.regions) > 1:

            # Create the columns of inter-regional links as a multi-index of the
            # pairs of regions and the transmitted carriers
            indexer = pd.MultiIndex.from_product(
                [self.lines_list, self.glob_mapping["Carriers_glob"]["Carrier"]],
                names=["Line", "Transmitted Carrier"],
            )

            self.connection_sheet_ids = {
                "F_OM": {
                    "value": 0,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": indexer,
                },
                "V_OM": {
                    "value": 0,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": indexer,
                },
                "Residual_capacity": {
                    "value": 0,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": indexer,
                },
                "Capacity_factor_line": {
                    "value": 1,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": indexer,
                },
                "Line_efficiency": {
                    "value": 1,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": indexer,
                },
                "AnnualProd_perunit_capacity": {
                    "value": 1,
                    "index": pd.Index(
                        ["AnnualProd_Per_UnitCapacity"], name="Performance Parameter"
                    ),
                    "columns": indexer,
                },
            }

            self.global_sheet_ids = {
                "Max_production_global": {
                    "value": 1e30,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": self.glob_mapping["Technologies_glob"].loc[
                        (
                            self.glob_mapping["Technologies_glob"]["Tech_category"]
                            != "Demand"
                        )
                        & (
                            self.glob_mapping["Technologies_glob"]["Tech_category"]
                            != "Storage"
                        )
                    ]["Technology"],
                },
                "Min_production_global": {
                    "value": 0,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": self.glob_mapping["Technologies_glob"].loc[
                        (
                            self.glob_mapping["Technologies_glob"]["Tech_category"]
                            != "Demand"
                        )
                        & (
                            self.glob_mapping["Technologies_glob"]["Tech_category"]
                            != "Storage"
                        )
                    ]["Technology"],
                },
                "Glob_emission_cap_annual": {
                    "value": 1e30,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": ["Global Emission Cap"],
                },
            }

            connections_operation_sorted = sorted(self.connection_sheet_ids.items(), key=lambda pair: list_connection_operation.index(pair[0]))
            self.connection_sheet_ids_sorted = dict(connections_operation_sorted)

            if self.mode == "Planning":

                self.connection_sheet_ids.update(
                    {
                        "INV": {
                            "value": 0,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": indexer,
                        },
                        "Decom_cost": {
                            "value": 0,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": indexer,
                        },
                        "Min_totalcap": {
                            "value": 0,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": indexer,
                        },
                        "Max_totalcap": {
                            "value": 1e10,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": indexer,
                        },
                        "Min_newcap": {
                            "value": 0,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": indexer,
                        },
                        "Max_newcap": {
                            "value": 1e10,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": indexer,
                        },
                        "Line_lifetime": {
                            "value": 1,
                            "index": pd.Index(
                                ["Technical Life Time"], name="Performance Parameter"
                            ),
                            "columns": indexer,
                        },
                        "Line_Economic_life": {
                            "value": 1,
                            "index": pd.Index(
                                ["Economic Life time"], name="Performance Parameter"
                            ),
                            "columns": indexer,
                        },
                        "Interest_rate": {
                            "value": 0.05,
                            "index": pd.Index(
                                ["Interest Rate"], name="Performance Parameter"
                            ),
                            "columns": indexer,
                        },
                    }
                )
                connections_planning_sorted = sorted(self.connection_sheet_ids.items(), key=lambda pair: list_connection_planning.index(pair[0]))
                self.connection_sheet_ids_sorted = dict(connections_planning_sorted)

                self.global_sheet_ids.update(
                    {
                        "Min_totalcap_global": {
                            "value": 0,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": self.glob_mapping["Technologies_glob"].loc[
                                self.glob_mapping["Technologies_glob"]["Tech_category"]
                                != "Demand"
                            ]["Technology"],
                        },
                        "Max_totalcap_global": {
                            "value": 1e10,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": self.glob_mapping["Technologies_glob"].loc[
                                self.glob_mapping["Technologies_glob"]["Tech_category"]
                                != "Demand"
                            ]["Technology"],
                        },
                        "Min_newcap_global": {
                            "value": 0,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": self.glob_mapping["Technologies_glob"].loc[
                                self.glob_mapping["Technologies_glob"]["Tech_category"]
                                != "Demand"
                            ]["Technology"],
                        },
                        "Max_newcap_global": {
                            "value": 1e10,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": self.glob_mapping["Technologies_glob"].loc[
                                self.glob_mapping["Technologies_glob"]["Tech_category"]
                                != "Demand"
                            ]["Technology"],
                        },
                        "Discount_rate": {
                            "value": 0.05,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": ["Annual Discount Rate"],
                        },
                    }
                )

        take_sorted_sheets= take_regional_sheets(self.mode,
        self.Technologies,self.regions)
        self.regional_sheets_ids = {}
        self.regional_sheets_ids_sorted= {}
        indexer_reg = {}
        indexer_reg_drop1 = {}
        indexer_reg_drop2 = {}
        add_indexer = {}
        conversion_plus_indexin = {}
        conversion_plus_indexout = {}

        # Creates the columns of the carrier_ratio_in and carrier_ratio_out sheets
        # by finding the conversion plus technologies and their input and output carriers
        for reg in self.regions:

            if "Conversion_plus" in self.Technologies[reg].keys():

                take_carrierin = [
                    self.mapping[reg]["Carrier_input"]
                    .loc[self.mapping[reg]["Carrier_input"]["Technology"] == tech][
                        "Carrier_in"
                    ]
                    .values
                    for tech in self.Technologies[reg]["Conversion_plus"]
                ]

                take_carrierin_ = [
                    carr
                    for index, value in enumerate(take_carrierin)
                    for carr in take_carrierin[index]
                ]

                take_technologyin = [
                    self.mapping[reg]["Carrier_input"]
                    .loc[self.mapping[reg]["Carrier_input"]["Technology"] == tech][
                        "Technology"
                    ]
                    .values
                    for tech in self.Technologies[reg]["Conversion_plus"]
                ]

                take_technologyin_ = [
                    tech
                    for index, value in enumerate(take_technologyin)
                    for tech in take_technologyin[index]
                ]

                take_carrierout = [
                    self.mapping[reg]["Carrier_output"]
                    .loc[self.mapping[reg]["Carrier_output"]["Technology"] == tech][
                        "Carrier_out"
                    ]
                    .values
                    for tech in self.Technologies[reg]["Conversion_plus"]
                ]

                take_carrierout_ = [
                    carr
                    for index, value in enumerate(take_carrierout)
                    for carr in take_carrierout[index]
                ]

                take_technologyout = [
                    self.mapping[reg]["Carrier_output"]
                    .loc[self.mapping[reg]["Carrier_output"]["Technology"] == tech][
                        "Technology"
                    ]
                    .values
                    for tech in self.Technologies[reg]["Conversion_plus"]
                ]

                take_technologyout_ = [
                    tech
                    for index, value in enumerate(take_technologyout)
                    for tech in take_technologyout[index]
                ]

                conversion_plus_indexin[reg] = pd.MultiIndex.from_arrays(
                    [take_technologyin_, take_carrierin_],
                    names=["Tech_category", "Technology"],
                )
                conversion_plus_indexout[reg] = pd.MultiIndex.from_arrays(
                    [take_technologyout_, take_carrierout_],
                    names=["Tech_category", "Technology"],
                )

            # Creates the columns of the technology-specific parameter files
            # based on the technology categories and the technologies within each
            # caregory
            dict_ = self.Technologies[reg]
            level1 = []
            level2 = []
            for key, values in dict_.items():
                if key != "Demand":
                    for value in values:
                        level1.append(key)
                        level2.append(value)

            indexer_reg[reg] = pd.MultiIndex.from_arrays(
                [level1, level2], names=["Tech_category", "Technology"]
            )

            if "Storage" in self.Technologies[reg].keys():

                indexer_reg_drop1[reg] = indexer_reg[reg].drop("Storage", level=0)

            else:

                indexer_reg_drop1[reg] = indexer_reg[reg]

            if "Transmission" in self.Technologies[reg].keys():

                indexer_reg_drop2[reg] = indexer_reg_drop1[reg].drop(
                    "Transmission", level=0
                )

            else:

                indexer_reg_drop2[reg] = indexer_reg_drop1[reg]

            level1_ = level1 * 2
            level2_ = level2 * 2
            tax = []
            sub = []
            for tech in level2:
                tax.append("Tax")
                sub.append("Sub")
            taxsub = tax + sub
            add_indexer[reg] = pd.MultiIndex.from_arrays(
                [taxsub, level1_, level2_],
                names=["Taxes or Subsidies", "Tech_category", "Technology"],
            )

            self.regional_sheets_ids[reg] = {
                "F_OM": {
                    "value": 0,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": indexer_reg[reg],
                },
                "V_OM": {
                    "value": 0,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": indexer_reg[reg],
                },
                "Residual_capacity": {
                    "value": 0,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": indexer_reg[reg],
                },
                "Max_production": {
                    "value": 1e20,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": indexer_reg_drop2[reg],
                },
                "Min_production": {
                    "value": 0,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": indexer_reg_drop2[reg],
                },
                "Capacity_factor_tech": {
                    "value": 1,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": indexer_reg[reg],
                },
                "Tech_efficiency": {
                    "value": 1,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": indexer_reg_drop1[reg],
                },
                "Specific_emission": {
                    "value": 0,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": indexer_reg_drop2[reg],
                },
                "Carbon_tax": {
                    "value": 0,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": indexer_reg_drop2[reg],
                },
                "Fix_taxsub": {
                    "value": 0,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": add_indexer[reg],
                },
                "Emission_cap_annual": {
                    "value": 1e10,
                    "index": pd.Index(self.main_years, name="Years"),
                    "columns": ["Emission Cap"],
                },
                "AnnualProd_perunit_capacity": {
                    "value": 1,
                    "index": pd.Index(
                        ["AnnualProd_Per_UnitCapacity"], name="Performance Parameter"
                    ),
                    "columns": indexer_reg[reg],
                },
            }

            if self.mode == "Planning":

                self.regional_sheets_ids[reg].update(
                    {
                        "INV": {
                            "value": 0,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": indexer_reg[reg],
                        },
                        "Investment_taxsub": {
                            "value": 0,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": add_indexer[reg],
                        },
                        "Decom_cost": {
                            "value": 0,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": indexer_reg[reg],
                        },
                        "Min_totalcap": {
                            "value": 0,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": indexer_reg[reg],
                        },
                        "Max_totalcap": {
                            "value": 1e10,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": indexer_reg[reg],
                        },
                        "Min_newcap": {
                            "value": 0,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": indexer_reg[reg],
                        },
                        "Max_newcap": {
                            "value": 1e10,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": indexer_reg[reg],
                        },
                        "Discount_rate": {
                            "value": 0.05,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": ["Annual Discount Rate"],
                        },
                        "Tech_lifetime": {
                            "value": 1,
                            "index": pd.Index(
                                ["Technical Life time"], name="Performance Parameter"
                            ),
                            "columns": indexer_reg[reg],
                        },
                        "Economic_lifetime": {
                            "value": 1,
                            "index": pd.Index(
                                ["Economic Life time"], name="Performance Parameter"
                            ),
                            "columns": indexer_reg[reg],
                        },
                        "Interest_rate": {
                            "value": 0.05,
                            "index": pd.Index(
                                ["Interest Rate"], name="Performance Parameter"
                            ),
                            "columns": indexer_reg[reg],
                        },
                    }
                )

            if "Storage" in self.Technologies[reg].keys():

                self.regional_sheets_ids[reg].update(
                    {
                        "Storage_charge_efficiency": {
                            "value": 1,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": pd.Index(
                                self.Technologies[reg]["Storage"], name="Technology"
                            ),
                        },
                        "Storage_discharge_efficiency": {
                            "value": 1,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": pd.Index(
                                self.Technologies[reg]["Storage"], name="Technology"
                            ),
                        },
                        "Storage_min_SOC": {
                            "value": 0,
                            "index": pd.Index(self.main_years, name="Years"),
                            "columns": pd.Index(
                                self.Technologies[reg]["Storage"], name="Technology"
                            ),
                        },
                        "Storage_initial_SOC": {
                            "value": 0,
                            "index": ["Initial State of Charge"],
                            "columns": pd.Index(
                                self.Technologies[reg]["Storage"], name="Technology"
                            ),
                        },
                        "Storage_charge_time": {
                            "value": 1,
                            "index": ["Charging Time of Storage in Hours"],
                            "columns": pd.Index(
                                self.Technologies[reg]["Storage"], name="Technology"
                            ),
                        },
                        "Storage_discharge_time": {
                            "value": 1,
                            "index": ["Discharging Time of Storage in Hours"],
                            "columns": pd.Index(
                                self.Technologies[reg]["Storage"], name="Technology"
                            ),
                        },
                    }
                )

            if "Timesteps" in self.glob_mapping.keys():

                self.time_steps = list(self.glob_mapping["Timesteps"]["Timeslice"])
                self.timeslice_fraction = self.glob_mapping["Timesteps"][
                    "Timeslice_fraction"
                ].values

                self.regional_sheets_ids[reg].update(
                    {
                        "Demand": {
                            "value": 0,
                            "index": pd.MultiIndex.from_product(
                                [self.main_years, self.time_steps],
                                names=["Years", "Timesteps"],
                            ),
                            "columns": self.Technologies[reg]["Demand"],
                        },
                        "capacity_factor_resource": {
                            "value": 1,
                            "index": pd.MultiIndex.from_product(
                                [self.main_years, self.time_steps],
                                names=["Years", "Timesteps"],
                            ),
                            "columns": indexer_reg_drop1[reg],
                        },
                        "Max_production_h": {
                            "value": 1e20,
                            "index": pd.MultiIndex.from_product(
                                [self.main_years, self.time_steps],
                                names=["Years", "Timesteps"],
                            ),
                            "columns": indexer_reg_drop2[reg],
                        },
                        "Min_production_h": {
                            "value": 0,
                            "index": pd.MultiIndex.from_product(
                                [self.main_years, self.time_steps],
                                names=["Years", "Timesteps"],
                            ),
                            "columns": indexer_reg_drop2[reg],
                        },
                    }
                )
                if "Conversion_plus" in self.Technologies[reg].keys():

                    self.regional_sheets_ids[reg].update(
                        {
                            "Carrier_ratio_in": {
                                "value": 1,
                                "index": pd.MultiIndex.from_product(
                                    [self.main_years, self.time_steps],
                                    names=["Years", "Timesteps"],
                                ),
                                "columns": conversion_plus_indexin[reg],
                            },
                            "Carrier_ratio_out": {
                                "value": 1,
                                "index": pd.MultiIndex.from_product(
                                    [self.main_years, self.time_steps],
                                    names=["Years", "Timesteps"],
                                ),
                                "columns": conversion_plus_indexout[reg],
                            },
                        }
                    )

            else:

                self.time_steps = ["Annual"]
                self.timeslice_fraction = 1

                self.regional_sheets_ids[reg].update(
                    {
                        "Demand": {
                            "value": 0,
                            "index": pd.MultiIndex.from_arrays(
                                [self.main_years, self.main_years],
                                names=["Years", "Timesteps"],
                            ),
                            "columns": self.Technologies[reg]["Demand"],
                        },
                        "capacity_factor_resource": {
                            "value": 1,
                            "index": pd.MultiIndex.from_arrays(
                                [self.main_years, self.main_years],
                                names=["Years", "Timesteps"],
                            ),
                            "columns": indexer_reg_drop1[reg],
                        },
                        "Max_production_h": {
                            "value": 1e20,
                            "index": pd.MultiIndex.from_product(
                                [self.main_years, self.main_years],
                                names=["Years", "Timesteps"],
                            ),
                            "columns": indexer_reg_drop2[reg],
                        },
                        "Min_production_h": {
                            "value": 0,
                            "index": pd.MultiIndex.from_product(
                                [self.main_years, self.main_years],
                                names=["Years", "Timesteps"],
                            ),
                            "columns": indexer_reg_drop2[reg],
                        },
                    }
                )

                if "Conversion_plus" in self.Technologies[reg].keys():

                    self.regional_sheets_ids[reg].update(
                        {
                            "Carrier_ratio_in": {
                                "value": 1,
                                "index": pd.MultiIndex.from_arrays(
                                    [self.main_years, self.main_years],
                                    names=["Years", "Timesteps"],
                                ),
                                "columns": conversion_plus_indexin[reg],
                            },
                            "Carrier_ratio_out": {
                                "value": 1,
                                "index": pd.MultiIndex.from_arrays(
                                    [self.main_years, self.main_years],
                                    names=["Years", "Timesteps"],
                                ),
                                "columns": conversion_plus_indexout[reg],
                            },
                        }
                    )
            
            regional_sheets_sorted = sorted(self.regional_sheets_ids[reg].items(), key=lambda pair: take_sorted_sheets[reg].index(pair[0]))
            self.regional_sheets_ids_sorted[reg] = dict(regional_sheets_sorted)
            

    def _write_input_excel(self, path):
        """
        Creates and writes the parameter files with their default values
        """

        if len(self.regions) > 1:

            with pd.ExcelWriter(
                r"{}/parameters_connections.xlsx".format(path)
            ) as writer:

                for key, value in self.connection_sheet_ids_sorted.items():

                    connection_data = pd.DataFrame(
                        value["value"], index=value["index"], columns=value["columns"]
                    )
                    connection_data.to_excel(writer, sheet_name=key)

            with pd.ExcelWriter(r"{}/parameters_global.xlsx".format(path)) as writer:

                for key, value in self.global_sheet_ids.items():

                    global_data = pd.DataFrame(
                        value["value"], index=value["index"], columns=value["columns"]
                    )
                    global_data.to_excel(writer, sheet_name=key)

        for reg in self.regions:

            with pd.ExcelWriter(r"{}/parameters_{}.xlsx".format(path, reg)) as writer:

                for key, value in self.regional_sheets_ids_sorted[reg].items():

                    regional_data = pd.DataFrame(
                        value["value"], index=value["index"], columns=value["columns"]
                    )
                    regional_data.to_excel(writer, sheet_name=key)

    def _read_data(self, path):

        """
        Reads the parameters with the given values by the user
        """

        if len(self.regions) > 1:

            trade_data_ids = take_trade_ids(mode=self.mode)
            global_data_ids = take_global_ids(mode=self.mode)
            check_sheet_name(path, "parameters_connections", trade_data_ids)
            check_sheet_name(path, "parameters_global", global_data_ids)

            trade_data = {}
            for key, value in trade_data_ids.items():

                trade_data[key] = pd.read_excel(
                    r"{}/parameters_connections.xlsx".format(path),
                    sheet_name=value["sheet_name"],
                    index_col=value["index_col"],
                    header=value["header"],
                )
                check_nan(
                    value["sheet_name"], trade_data[key], "parameters_connections"
                )
                check_index_data(
                    trade_data[key].index,
                    value["sheet_name"],
                    "parameters_connections",
                    pd.Index(self.connection_sheet_ids[value["sheet_name"]]["index"]),
                )

                check_index_data(
                    trade_data[key].columns,
                    value["sheet_name"],
                    "parameters_connections",
                    pd.Index(self.connection_sheet_ids[value["sheet_name"]]["columns"]),
                )

            global_data = {}
            for key, value in global_data_ids.items():
                global_data[key] = pd.read_excel(
                    r"{}/parameters_global.xlsx".format(path),
                    sheet_name=value["sheet_name"],
                    index_col=value["index_col"],
                    header=value["header"],
                )
                check_nan(value["sheet_name"], global_data[key], "parameters_global")

                check_index_data(
                    global_data[key].index,
                    value["sheet_name"],
                    "parameters_global",
                    pd.Index(self.global_sheet_ids[value["sheet_name"]]["index"]),
                )

                check_index_data(
                    global_data[key].columns,
                    value["sheet_name"],
                    "parameters_global",
                    pd.Index(self.global_sheet_ids[value["sheet_name"]]["columns"]),
                )

            self.trade_data = trade_data
            self.global_data = global_data

        data = {}
        ids = take_ids(self.regions, self.Technologies, self.mode)
        for reg in self.regions:

            check_sheet_name(path, "parameters_{}".format(reg), ids[reg])
            reg_data = {}
            for key, value in ids[reg].items():

                reg_data[key] = pd.read_excel(
                    r"{}/parameters_{}.xlsx".format(path, reg),
                    sheet_name=value["sheet_name"],
                    index_col=value["index_col"],
                    header=value["header"],
                )
                check_nan(
                    value["sheet_name"], reg_data[key], "parameters_{}".format(reg)
                )

                check_index_data(
                    reg_data[key].index,
                    value["sheet_name"],
                    "parameters_{}.xlsx".format(reg),
                    pd.Index(
                        self.regional_sheets_ids[reg][value["sheet_name"]]["index"]
                    ),
                )

                check_index_data(
                    reg_data[key].columns,
                    value["sheet_name"],
                    "parameters_{}.xlsx".format(reg),
                    pd.Index(
                        self.regional_sheets_ids[reg][value["sheet_name"]]["columns"]
                    ),
                )

            data[reg] = reg_data

        self.data = data

    @property
    def multi_node(self):
        if len(self.regions) > 1:
            return True
        return False

    @property
    def mode(self):
        return self._mode

    @mode.setter
    def mode(self, var):
        if var not in MODES:

            raise WrongInputMode(
                f"The given optimization mode is not valid.\
                                 Valide modes are {MODES}"
            )

        else:

            self._mode = var
